"""
 * Copyright (c) KylinSoft  Co., Ltd. 2024.All rights reserved.
 * PilotGo-plugin licensed under the Mulan Permissive Software License, Version 2. 
 * See LICENSE file for more details.
 * Author: wangqingzheng <wangqingzheng@kylinos.cn>
 * Date: Fri Feb 23 10:00:37 2024 +0800
"""
#!/usr/bin/env python
# encoding: utf-8
"""
公共函数
@author: Wqz
@time: 11/6/19 4:33 PM
"""
from django.core.paginator import Paginator, EmptyPage
from django.forms import model_to_dict
from django.http import JsonResponse
from rest_framework import pagination, status
from rest_framework.permissions import BasePermission


def return_time(test_time):
    time = test_time.split('-')[0] + '-' + test_time.split('-')[1] + '-' + test_time.split('-')[2] + ' ' + \
               test_time.split('-')[3] + ':' + test_time.split('-')[4] + ':' + test_time.split('-')[5]
    return time


def json_response(data=None, code=None, message=None):
    """
    返回自定义格式数据
    :param data:
    :param code:
    :param message:
    :return:
    """
    res = {
        'data': data,
        'code': code,
        'message': message
    }
    return JsonResponse(res)


def list_response(result, code, message):
    """
    :param result:
    :param code:
    :param message:
    :return:
    """
    res = {}
    if result.data:
        res['data'] = result.data
    if code:
        res['code'] = code
    if message:
        res['message'] = message
    return JsonResponse(res)


def model_to_dict_myself(queryset, **kwargs):
    """
        返回model_to_dict转换的字段
        """
    if not queryset:
        return {}
    data = []
    if 'exclude' in kwargs.keys():
        for query in queryset:
            dict = model_to_dict(query, exclude=kwargs['exclude'])
            data.append(dict)
    elif 'fields' in kwargs.keys():
        for query in queryset:
            dict = model_to_dict(query, fields=kwargs['fields'])
            data.append(dict)
    else:
        for query in queryset:
            dict = model_to_dict(query)
            data.append(dict)
    return data


def jwt_response_payload_handler(token, user=None, request=None):
    """
    自定义jwt认证成功返回数据
    """
    return {
        'token': token,
        'user_id': user.id,
        'username': user.username,
        'is_superuser': user.is_superuser,
        'is_staff': user.is_staff,
        'code': 200,
    }

def get_error_message(serializer):
    """
    返回错误信息
    :param serializer:
    :return:
    """
    for _, error in serializer.errors.items():
        return error[0]


class LimsPageSet(pagination.PageNumberPagination):
    """
    分页设置
    分页样式  ?page=1&page_size=10
    """
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000
    page_query_param = 'page'


# class ZbmPermission(BasePermission):
#     """
#     管理员和超级管理员可以对系统进行各种
#     权限操作，普通用户只能对信息进行查看
#     """
#     def has_permission(self, request, view):
#         try:
#             user_type = request.user.user_type_choices
#             if user_type == 2 or request.user.is_superuser or request.method == 'GET':
#                 return True
#             return False
#         except:
#             pass


def get_page(data, objs):
    """
    进行分页
    :param data:
    :param objs:
    :return:
    """
    try:
        page = int(data.get('page', 1))
        page_size = int(data.get('page_size', 5))
    except Exception as e:
        return json_response({}, status.HTTP_400_BAD_REQUEST, '参数类型不对')
    paginator = Paginator(objs, page_size)  # 设置每一页显示几条  创建一个panginator对象
    try:
        current_num = page  # 当在url内输入的?page = 页码数  显示你输入的页面数目 默认为第2页
        list = paginator.page(current_num)
    except EmptyPage:
        list = paginator.page(1)  # 当输入的page是不存在的时候就会报错
    return list




import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class Config():
    def __init__(self):
        # 定义样式
        self.thin = Side(border_style="thin", color="000000")
        self.border_thin = Border(top=self.thin, left=self.thin, right=self.thin, bottom=self.thin)

        # 创建一个居中对齐的样式对象
        # self.alignment_center = Alignment(horizontal='center', vertical='center')
        # 居中显示
        self.alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
        self.alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)

        # (1,1)的位置
        self.color_title = PatternFill("solid", fgColor="CC99FF")  # (1,1)的颜色
        self.font_title = Font(name="宋体", size=20, bold=True, color="993300")

        # 表头的其它颜色
        self.color_col_top = PatternFill("solid", fgColor="99CCFF")
        self.font_col_top = Font(name="宋体", size=18, color="000000")

        # cmd的颜色
        self.color_cmd = PatternFill("solid", fgColor="CCFFFF")
        self.font_cmd = Font(name="宋体", size=13, bold=True, color="000000")

        # 左侧第一列的颜色->黄色
        self.color_item_1 = PatternFill("solid", fgColor="FFCC99")
        self.font_item_1 = Font(name="宋体", size=13, color="000000")

        # 左侧第二列的颜色->绿色
        self.color_item_2 = PatternFill("solid", fgColor="CCFFCC")
        self.font_item_2 = Font(name="宋体", size=11, color="000000")

        # 普通
        self.color_data = PatternFill()
        self.font_data = Font(name="宋体")

        # 对比下降内容
        self.color_compar_up = PatternFill("solid", fgColor="c6efce")
        self.font_compar_up = Font(name="宋体", color="006100")

        # 对比上升内容
        self.color_compar_down = PatternFill("solid", fgColor="ffc7ce")
        self.font_compar_down = Font(name="宋体",color="9c0006")

def set_cell_style(worksheet, row, column,color, font):
    """
    设置指定工作表中指定单元格的样式。
    Args:
        worksheet: 工作表对象。
        row: 需要改变的单元格的行
        column: 需要改变的单元格的列
        color: 单元格的填充样式。
        font: 单元格的字体样式。
    """
    cell = worksheet.cell(row=row, column=column)
    cell.fill = color
    cell.font = font

def stream_excel(data):
    config = Config()
    # 创建一个工作簿对象
    workbook = openpyxl.Workbook()
    # 获取默认的工作表对象
    worksheet = workbook.active
    # 获取表头
    header = list(data['data'][0].keys())
    worksheet.append(header)

    # 制作基础表格
    for row_data in data['data']:
        row = []
        for key in header:
            row.append(row_data.get(key, ''))
        worksheet.append(row)

    # 删除Column的表头
    worksheet.delete_rows(1)

    # 合并行
    worksheet.merge_cells('A1:B1')
    worksheet.merge_cells('A2:B2')
    worksheet.merge_cells('A3:B3')
    worksheet.merge_cells('A4:B4')
    # 合并列
    worksheet.merge_cells(start_row=5, start_column=1, end_row=9, end_column=1)
    worksheet.merge_cells(start_row=10, start_column=1, end_row=14, end_column=1)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1

    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置表头样式
    for cell in worksheet[1]:
        cell.border = config.border_thin
        cell.fill = config.color_col_top
        cell.font = config.font_col_top

    # 设置(1,1)的颜色
    set_cell_style(worksheet, 1, 1, config.color_title, config.font_title)

    # 设置项目名称,cmd,修改参数的颜色
    for i in range(2,5):
        set_cell_style(worksheet, i, 1, config.color_cmd, config.font_cmd)

    # 遍历每个单元格，设置列宽,行高,居中对齐,和对比值的特定颜色
    for i in range(1, worksheet.max_column + 1):
        # 设置列宽
        column_letter = get_column_letter(i)
        if i == 1:
            worksheet.column_dimensions[column_letter].width = 10
        else:
            worksheet.column_dimensions[column_letter].width = 20
        # 如果是对比值列，则遍历每个单元格
        if worksheet.cell(row=1, column=i).value == '对比值':
            for row in worksheet.iter_rows(min_row=2, min_col=i, max_col=i):
                cell = row[0]
                cell_value = cell.value
                if cell_value and float(cell_value[:-1]) > 10:
                    cell.fill = config.color_compar_up
                    cell.font = config.font_compar_up
                elif cell_value and float(cell_value[:-1]) < -10:
                    cell.fill = config.color_compar_down
                    cell.font = config.font_compar_down
                print(f"单元格({row[0].row}, {column_letter}): {cell_value}")
        # 设置行高
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = config.border_thin
                cell.alignment = config.alignment_center
            if row[0].row == 1 or row[0].row == 4:
                worksheet.row_dimensions[row[0].row].height = 50
            elif row[0].row == 2 or row[0].row == 3:
                worksheet.row_dimensions[row[0].row].height = 25
            else:
                worksheet.row_dimensions[row[0].row].height = 13

    # 保存工作簿到文件
    workbook.save('mydata.xlsx')
