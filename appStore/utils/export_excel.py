"""
 * Copyright (c) KylinSoft  Co., Ltd. 2024.All rights reserved.
 * PilotGo-plugin licensed under the Mulan Permissive Software License, Version 2.
 * See LICENSE file for more details.
 * Author: wangqingzheng <wangqingzheng@kylinos.cn>
 * Date: Fri Feb 23 11:15:41 2024 +0800
"""
#!/usr/bin/env python
# encoding: utf-8
"""
导出表格
@author: Wqz
@time: 26/7/23 4:33 PM
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class Config():
    def __init__(self):
        # 定义样式
        self.thin = Side(border_style="thin", color="000000")
        self.border_thin = Border(top=self.thin, left=self.thin, right=self.thin, bottom=self.thin)

        # 创建一个居中对齐的样式对象
        # self.alignment_center = Alignment(horizontal='center', vertical='center')
        # 居中显示
        self.alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
        self.alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)

        # (1,1)的位置
        self.color_title = PatternFill("solid", fgColor="CC99FF")  # (1,1)的颜色
        self.font_title = Font(name="宋体", size=20, bold=True, color="993300")

        # 表头的其它颜色
        self.color_col_top = PatternFill("solid", fgColor="99CCFF")
        self. font_col_top = Font(name="宋体", size=18, color="000000")

        # cmd的颜色
        self.color_cmd = PatternFill("solid", fgColor="CCFFFF")
        self.font_cmd = Font(name="宋体", size=13, bold=True, color="000000")

        # 左侧第一列的颜色->黄色
        self.color_item_1 = PatternFill("solid", fgColor="FFCC99")
        self.font_item_1 = Font(name="宋体", size=13, color="000000")

        # 左侧第二列的颜色->绿色
        self.color_item_2 = PatternFill("solid", fgColor="CCFFCC")
        self.font_item_2 = Font(name="宋体", size=11, color="000000")

        # 普通
        self.color_data = PatternFill()
        self.font_data = Font(name="宋体")

        # 对比下降内容
        self.color_compar_up = PatternFill("solid", fgColor="c6efce")
        self.font_compar_up = Font(name="宋体", color="006100")

        # 对比上升内容
        self.color_compar_down = PatternFill("solid", fgColor="ffc7ce")
        self.font_compar_down = Font(name="宋体",color="9c0006")
        # 左侧的列宽，按照顺序计算
        self.sheetname_width = {"Stream": [10, 20], "Lmbench": [30, 20], "Unixbench": [10, 30], "Fio": [15, 10],
                                "Iozone": [30, 10], "Specjvm2008": [10, 20], "Speccpu2006(base)": [10,20],
                                "Speccpu2006(peak)": [10, 20], "Speccpu2017(base)": [10, 20],
                                "Speccpu2017(peak)": [10, 20]}

# 参数是{单元格位置(1,1),填充颜色(color),字体样式(font)}
def set_cell_style(worksheet, row, column,color, font):
    """
    设置指定工作表中指定单元格的样式。
    Args:
        worksheet: 工作表对象。
        row: 需要改变的单元格的行
        column: 需要改变的单元格的列
        alg: 单元格的对齐方式，类型为 openpyxl.styles.Alignment 对象。
        color: 单元格的填充样式。
        font: 单元格的字体样式。
    """
    cell = worksheet.cell(row=row, column=column)
    cell.fill = color
    cell.font = font

def create_base_export(sheetname,data):
    if os.path.exists('./mydata.xlsx'):
        print(sheetname,'数据===文件已经存在，直接打开')
        workbook = openpyxl.open('./mydata.xlsx')
    else:
        print('文件不存在，新建这个文件')
        workbook = openpyxl.Workbook()
    # 获取默认的工作表对象
    worksheet = workbook.active

    if sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
    else:
        worksheet = workbook.create_sheet(sheetname)

    # 获取表头
    header = list(data['data'][0].keys())
    worksheet.append(header)

    # 制作基础表格
    for row_data in data['data']:
        row = []
        for key in header:
            row.append(row_data.get(key, ''))
        worksheet.append(row)

    # 删除Column的表头
    worksheet.delete_rows(1)

    # 合并表头中的行
    if sheetname in ["Stream", "Lmbench","Unixbench","Fio","Specjvm2008"]:
        worksheet.merge_cells('A1:B1')
        worksheet.merge_cells('A2:B2')
        worksheet.merge_cells('A3:B3')
        worksheet.merge_cells('A4:B4')
    elif sheetname in ["Speccpu2006(base)","Speccpu2006(peak)"]:
        worksheet.merge_cells('A1:D1')
        worksheet.merge_cells('A2:D2')
        worksheet.merge_cells('A3:D3')
        worksheet.merge_cells('A4:D4')
    elif sheetname in ["Speccpu2017(base)","Speccpu2017(peak)"]:
        worksheet.merge_cells('A1:E1')
        worksheet.merge_cells('A2:E2')
        worksheet.merge_cells('A3:E3')
        worksheet.merge_cells('A4:E4')
    return worksheet,workbook

def set_base_export_style(worksheet,config,sheetname):
    # 设置表头样式
    for cell in worksheet[1]:
        cell.fill = config.color_col_top
        cell.font = config.font_col_top

    # 设置(1,1)的颜色
    set_cell_style(worksheet, 1, 1, config.color_title, config.font_title)

    # 设置项目名称,cmd,修改参数的颜色
    for i in range(2, 5):
        set_cell_style(worksheet, i, 1, config.color_cmd, config.font_cmd)

    # 遍历每个单元格，设置列宽,行高,居中对齐,和对比值的特定颜色
    for i in range(1, worksheet.max_column + 1):
        # 设置列宽
        column_letter = get_column_letter(i)
        if sheetname in ["Speccpu2006(base)", "Speccpu2006(peak)"]:
            if i in [1,2,3]:
                worksheet.column_dimensions[column_letter].width = config.sheetname_width[sheetname][0]
            elif i == 4:
                worksheet.column_dimensions[column_letter].width = config.sheetname_width[sheetname][1]
            else:
                worksheet.column_dimensions[column_letter].width = 15
        elif sheetname in ["Speccpu2017(base)", "Speccpu2017(peak)"]:
            if i in [1,2,3,4]:
                worksheet.column_dimensions[column_letter].width = config.sheetname_width[sheetname][0]
            elif i == 5:
                worksheet.column_dimensions[column_letter].width = config.sheetname_width[sheetname][1]
            else:
                worksheet.column_dimensions[column_letter].width = 15
        else:
            if i == 1:
                worksheet.column_dimensions[column_letter].width = config.sheetname_width[sheetname][0]
            elif i == 2:
                worksheet.column_dimensions[column_letter].width = config.sheetname_width[sheetname][1]
            else:
                worksheet.column_dimensions[column_letter].width = 15
        # 如果是对比值列，则遍历每个单元格
        if worksheet.cell(row=1, column=i).value == '对比值':
            for row in worksheet.iter_rows(min_row=2, min_col=i, max_col=i):
                cell = row[0]
                cell_value = cell.value
                if cell_value and float(cell_value[:-1]) > 10:
                    cell.fill = config.color_compar_up
                    cell.font = config.font_compar_up
                elif cell_value and float(cell_value[:-1]) < -10:
                    cell.fill = config.color_compar_down
                    cell.font = config.font_compar_down
                # 每个数据
                # print(f"单元格({row[0].row}, {column_letter}): {cell_value}")
    # 设置行高
    worksheet.row_dimensions[1].height = 50
    worksheet.row_dimensions[2].height = 25
    worksheet.row_dimensions[3].height = 25
    worksheet.row_dimensions[4].height = 50
    # 设置边框
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = config.border_thin
            cell.alignment = config.alignment_center
        # if row[0].row == 1 or row[0].row == 4:
        #     worksheet.row_dimensions[row[0].row].height = 50
        # elif row[0].row == 2 or row[0].row == 3:
        #     worksheet.row_dimensions[row[0].row].height = 25
    return worksheet

def env_excel(data):
    ### todo 增加一个锁判断 #############
    if os.path.exists('./mydata.xlsx'):
        os.remove('./mydata.xlsx')
    ################################################
    config = Config()
    workbook = openpyxl.Workbook()
    # 获取默认的工作表对象
    worksheet = workbook.active

    worksheet.title = "性能测试环境"

    # 获取表头
    header = list(data['data']['data'][0].keys())
    worksheet.append(header)

    # 制作基础表格
    for row_data in data['data']['data']:
        row = []
        for key in header:
            row.append(row_data.get(key, ''))
        worksheet.append(row)
    worksheet.delete_rows(1,2)
    worksheet.insert_rows(1, 5)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter][5:]:
        cell.fill = config.color_title
        cell.font = config.font_title
    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter][5:]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1
    # 设置第三列的样式
    column_letter = get_column_letter(3)
    for cell in worksheet[column_letter][5:]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置表头样式
    for cell in worksheet[1]:
        cell.fill = config.color_title
        cell.font = config.font_title

    # 设置宽度
    worksheet.column_dimensions["A"].width = 15
    worksheet.column_dimensions["B"].width = 15
    worksheet.column_dimensions["C"].width = 15
    worksheet.column_dimensions["D"].width = 50

    # 磁盘和网卡占位
    disk_row_nu = 0
    nic_row_nu = 0
    for i in range(1, worksheet.max_row + 1):
        # 如果是对比值列，则遍历每个单元格
        if worksheet.cell(row=i, column=2).value == 'disk':
            disk_row_nu += 1
        if worksheet.cell(row=i, column=2).value == 'nicinfo':
            nic_row_nu += 1
    d_n_row_nu = disk_row_nu + nic_row_nu
    # 合并列
    # 第一列
    worksheet.merge_cells(start_row=6, start_column=1, end_row=35+d_n_row_nu, end_column=1)
    worksheet.merge_cells(start_row=36+d_n_row_nu, start_column=1, end_row=worksheet.max_row, end_column=1)
    # 第二列
    worksheet.merge_cells(start_row=6, start_column=2, end_row=8, end_column=2)
    worksheet.merge_cells(start_row=9, start_column=2, end_row=10, end_column=2)
    worksheet.merge_cells(start_row=11, start_column=2, end_row=27, end_column=2)
    worksheet.merge_cells(start_row=28, start_column=2, end_row=35, end_column=2)
    # disk的数据
    worksheet.merge_cells(start_row=36, start_column=2, end_row=35 + disk_row_nu, end_column=2)
    # nick的数据
    worksheet.merge_cells(start_row=36 + disk_row_nu, start_column=2, end_row=35 + d_n_row_nu, end_column=2)
    # os
    worksheet.merge_cells(start_row=36 + d_n_row_nu, start_column=2, end_row=41 + d_n_row_nu, end_column=2)
    # runtime
    worksheet.merge_cells(start_row=42 + d_n_row_nu, start_column=2, end_row=52 + d_n_row_nu, end_column=2)
    # software_ver
    worksheet.merge_cells(start_row=53 + d_n_row_nu, start_column=2, end_row=58 + d_n_row_nu, end_column=2)
    # nic
    worksheet.merge_cells(start_row=59 + d_n_row_nu, start_column=2, end_row=worksheet.max_row, end_column=2)

    # 合并行
    worksheet.merge_cells('A1:D1')
    # 设置行高
    worksheet.row_dimensions[1].height = 30
    # 设置边框
    for row in worksheet.iter_rows():
        if row[0].row == 1 or row[0].row > 5:
            for cell in row:
                cell.border = config.border_thin
                cell.alignment = config.alignment_center
        # if row[0].row == 1 or row[0].row == 4:
        #     worksheet.row_dimensions[row[0].row].height = 50
        # elif row[0].row == 2 or row[0].row == 3:
        #     worksheet.row_dimensions[row[0].row].height = 25

    # D列左对齐
    for row in worksheet["D"]:
        row.alignment = config.alignment_left

    worksheet.cell(row=1, column=1).value = "性能测试环境统计表"
    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

def stream_excel(data):
    sheetname = "Stream"
    config = Config()
    worksheet,workbook = create_base_export(sheetname,data)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1

    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置其它的表格样式
    worksheet = set_base_export_style(worksheet,config,sheetname)

    # 合并列
    worksheet.merge_cells(start_row=5, start_column=1, end_row=9, end_column=1)
    worksheet.merge_cells(start_row=10, start_column=1, end_row=14, end_column=1)
    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

def lmbench_excel(data):
    sheetname = "Lmbench"
    config = Config()

    worksheet, workbook = create_base_export(sheetname, data)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1

    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置其它的表格样式
    worksheet = set_base_export_style(worksheet, config,sheetname)

    # 合并列
    worksheet.merge_cells(start_row=5, start_column=1, end_row=9, end_column=1)
    worksheet.merge_cells(start_row=10, start_column=1, end_row=20, end_column=1)
    worksheet.merge_cells(start_row=21, start_column=1, end_row=25, end_column=1)
    worksheet.merge_cells(start_row=26, start_column=1, end_row=30, end_column=1)
    worksheet.merge_cells(start_row=31, start_column=1, end_row=34, end_column=1)
    worksheet.merge_cells(start_row=35, start_column=1, end_row=38, end_column=1)
    worksheet.merge_cells(start_row=39, start_column=1, end_row=45, end_column=1)
    worksheet.merge_cells(start_row=46, start_column=1, end_row=53, end_column=1)
    worksheet.merge_cells(start_row=54, start_column=1, end_row=61, end_column=1)
    worksheet.merge_cells(start_row=62, start_column=1, end_row=70, end_column=1)
    worksheet.merge_cells(start_row=71, start_column=1, end_row=75, end_column=1)
    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

def unixbench_excel(data):
    sheetname = "Unixbench"
    config = Config()
    worksheet, workbook = create_base_export(sheetname, data)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1

    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置其它的表格样式
    worksheet = set_base_export_style(worksheet, config,sheetname)

    # 合并列
    worksheet.merge_cells(start_row=5, start_column=1, end_row=17, end_column=1)
    worksheet.merge_cells(start_row=18, start_column=1, end_row=30, end_column=1)

    # B列左侧对齐
    for row in worksheet["B"]:
        row.alignment = config.alignment_left

    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

def fio_excel(data):
    sheetname = "Fio"
    config = Config()
    worksheet, workbook = create_base_export(sheetname, data)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1

    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置其它的表格样式
    worksheet = set_base_export_style(worksheet, config,sheetname)

    # 合并列
    bw_row = []
    for i in range(int((worksheet.max_row-4)/4)):
        bw_row.append(4*i + 8)
        worksheet.merge_cells(start_row=5+4*i, start_column=1, end_row=8+4*i, end_column=1)

    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

def iozone_excel(data):
    sheetname = "Iozone"
    config = Config()
    worksheet, workbook = create_base_export(sheetname, data)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1

    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置其它的表格样式
    worksheet = set_base_export_style(worksheet, config,sheetname)

    # 合并列

    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

def jvm2008_excel(data):
    sheetname = "Specjvm2008"
    config = Config()
    worksheet,workbook = create_base_export(sheetname,data)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1

    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置其它的表格样式
    worksheet = set_base_export_style(worksheet,config,sheetname)

    # 合并列
    worksheet.merge_cells(start_row=5, start_column=1, end_row=16, end_column=1)
    worksheet.merge_cells(start_row=17, start_column=1, end_row=28, end_column=1)

    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

def cpu2006_excel(sheetname, data):
    config = Config()

    worksheet,workbook = create_base_export(sheetname,data)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1
    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1
    # 设置第三列的样式
    column_letter = get_column_letter(3)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1
    # 设置第四列的样式
    column_letter = get_column_letter(4)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2

    # 设置其它的表格样式
    worksheet = set_base_export_style(worksheet,config,sheetname)

    # 合并列
    # 第一列
    worksheet.merge_cells(start_row=5, start_column=1, end_row=66, end_column=1)
    # 第二列
    worksheet.merge_cells(start_row=5, start_column=2, end_row=35, end_column=2)
    worksheet.merge_cells(start_row=36, start_column=2, end_row=66, end_column=2)
    # # 第三列
    worksheet.merge_cells(start_row=5, start_column=3, end_row=17, end_column=3)
    worksheet.merge_cells(start_row=18, start_column=3, end_row=35, end_column=3)
    worksheet.merge_cells(start_row=36, start_column=3, end_row=48, end_column=3)
    worksheet.merge_cells(start_row=49, start_column=3, end_row=66, end_column=3)
    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

def cpu2017_excel(sheetname, data):
    config = Config()
    worksheet,workbook = create_base_export(sheetname,data)

    # 设置第一列的样式
    column_letter = get_column_letter(1)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1
    # 设置第二列的样式
    column_letter = get_column_letter(2)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1
    # 设置第三列的样式
    column_letter = get_column_letter(3)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1
    # 设置第四列的样式
    column_letter = get_column_letter(4)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_1
        cell.font = config.font_item_1
    # 设置第五列的样式
    column_letter = get_column_letter(5)
    for cell in worksheet[column_letter]:
        cell.fill = config.color_item_2
        cell.font = config.font_item_2
    #
    # # 设置其它的表格样式
    worksheet = set_base_export_style(worksheet,config,sheetname)

    # 合并列
    # 第一列
    worksheet.merge_cells(start_row=5, start_column=1, end_row=54, end_column=1)
    # 第二列
    worksheet.merge_cells(start_row=5, start_column=2, end_row=29, end_column=2)
    worksheet.merge_cells(start_row=30, start_column=2, end_row=54, end_column=2)
    # 第三列
    worksheet.merge_cells(start_row=5, start_column=3, end_row=15, end_column=3)
    worksheet.merge_cells(start_row=16, start_column=3, end_row=29, end_column=3)
    worksheet.merge_cells(start_row=30, start_column=3, end_row=40, end_column=3)
    worksheet.merge_cells(start_row=41, start_column=3, end_row=54, end_column=3)
    # 第四列
    worksheet.merge_cells(start_row=5, start_column=4, end_row=54, end_column=4)
    # 保存工作簿到文件
    workbook.save('mydata.xlsx')

